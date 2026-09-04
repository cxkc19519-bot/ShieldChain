"""Durable local knowledge service backed by BGE, Milvus and local metadata."""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import tempfile
import time
from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from threading import RLock
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from uuid import UUID, uuid4, uuid5

from shieldchain.rag.answering import contains_prompt_injection
from shieldchain.rag.api_service import (
    KnowledgeInputRejected,
    KnowledgeNotFound,
    UploadedDocument,
)
from shieldchain.rag.evaluation import (
    EvaluationCase,
    EvaluationDataset,
    EvaluationObservation,
    load_evaluation_dataset,
)
from shieldchain.rag.evaluation import (
    evaluate as evaluate_dataset,
)
from shieldchain.rag.local_milvus import DEFAULT_COLLECTION, ensure_collection
from shieldchain.rag.local_semantic_chunking import (
    DeepSeekSemanticChunker,
    SemanticChunkingError,
    SemanticSegment,
)
from shieldchain.rag.schemas import (
    CitationView,
    CreateKnowledgeBaseRequest,
    DegradationView,
    DocumentChunkListResponse,
    DocumentVersionListResponse,
    DocumentVersionView,
    EvaluationCaseResultView,
    EvaluationRequest,
    EvaluationResponse,
    KnowledgeBaseDeleteResponse,
    KnowledgeBaseView,
    KnowledgeChunkView,
    KnowledgeDocumentListResponse,
    KnowledgeDocumentView,
    LifecycleOperationResponse,
    RetrievalHitView,
    RetrievalRequest,
    RetrievalResponse,
)

_TERM = re.compile(r"[a-z0-9_./:-]+|[\u4e00-\u9fff]", re.IGNORECASE)
_CJK_SEQUENCE = re.compile(r"[\u4e00-\u9fff]+")
_CLAIM_MARKER = re.compile(
    r"\[shieldchain-claim:([a-z0-9_.-]{1,64}):(supports|counters)\]",
    re.IGNORECASE,
)
_CHUNK_NAMESPACE = UUID("712c7cfa-f63d-4665-88bd-4d2edf3b5d1c")
_CHUNK_SIZE = 1_200
_CHUNK_OVERLAP = 180
_EVALUATION_K = 5
_QUALITY_THRESHOLDS = {
    "recall_at_k": 0.75,
    "mrr_at_k": 0.70,
    "ndcg_at_k": 0.70,
    "citation_correctness": 0.80,
    "citation_precision": 0.80,
    "expected_citation_recall": 0.90,
    "extractive_faithfulness": 1.00,
    "refusal_accuracy": 0.90,
}
_MAX_EVALUATION_FAILURE_RATE = 0.05
_EMBED_BATCH_SIZE = 64
_MIN_RERANK_RELEVANCE = 0.10
_RELATIVE_RERANK_RELEVANCE = 0.05
_GENERIC_TITLE_TERMS = {
    "安全",
    "事件",
    "分析",
    "报告",
    "技术",
    "数据",
    "网络",
    "管理",
}


@dataclass(frozen=True, slots=True)
class _LocalBm25Corpus:
    entries: tuple[tuple[dict[str, object], Counter[str], int], ...]
    average_length: float
    document_frequency: Counter[str]


class LocalRagUnavailable(RuntimeError):
    """The loopback model service or local Milvus instance cannot be used."""


class LocalKnowledgeService:
    """Persist uploaded files locally and index searchable chunks in local Milvus."""

    def __init__(
        self,
        root: str | Path | None = None,
        *,
        evaluation_root: str | Path | None = None,
    ) -> None:
        selected = (
            root or os.environ.get("SHIELDCHAIN_LOCAL_KNOWLEDGE_ROOT") or "data/local-knowledge"
        )
        self._root = Path(selected).expanduser().resolve()
        selected_evaluation = (
            evaluation_root
            or os.environ.get("RAG_EVALUATION_ROOT")
            or "sample_docs/security_vertical/evaluation"
        )
        self._evaluation_root = Path(selected_evaluation).expanduser().resolve()
        self._catalog_path = self._root / "catalog.json"
        self._model_url = os.environ.get("SHIELDCHAIN_LOCAL_RAG_URL", "http://127.0.0.1:8001")
        self._milvus_url = os.environ.get("SHIELDCHAIN_LOCAL_MILVUS_URI", "http://127.0.0.1:19530")
        self._collection = os.environ.get("SHIELDCHAIN_LOCAL_MILVUS_COLLECTION", DEFAULT_COLLECTION)
        self._lock = RLock()
        self._root.mkdir(parents=True, exist_ok=True)

    def list_knowledge_bases(self, *, tenant_id: UUID) -> Sequence[KnowledgeBaseView]:
        with self._lock:
            return [KnowledgeBaseView.model_validate(value) for value in self._catalog()["bases"]]

    def create_knowledge_base(
        self, payload: CreateKnowledgeBaseRequest, *, tenant_id: UUID
    ) -> KnowledgeBaseView:
        with self._lock:
            catalog = self._catalog()
            now = datetime.now(UTC)
            base = KnowledgeBaseView(
                id=uuid4(),
                name=payload.name,
                status="draft",
                default_sensitivity=payload.default_sensitivity,
                version_policy=payload.version_policy,
                created_at=now,
                updated_at=now,
            )
            catalog["bases"].append(base.model_dump(mode="json"))
            catalog["documents"][str(base.id)] = []
            self._save(catalog)
            return base

    def delete_knowledge_base(
        self, knowledge_base_id: UUID, *, tenant_id: UUID
    ) -> KnowledgeBaseDeleteResponse:
        with self._lock:
            catalog = self._catalog()
            records = list(self._records(catalog, knowledge_base_id))
            for record in records:
                document = KnowledgeDocumentView.model_validate(record["view"])
                for version in document.versions:
                    self._delete_vectors(version.id)
                shutil.rmtree(self._root / "documents" / str(document.id), ignore_errors=True)
            catalog["bases"] = [
                base for base in catalog["bases"] if base.get("id") != str(knowledge_base_id)
            ]
            catalog["documents"].pop(str(knowledge_base_id), None)
            self._save(catalog)
            return KnowledgeBaseDeleteResponse(id=knowledge_base_id, status="completed")

    def upload_document(
        self, knowledge_base_id: UUID, upload: UploadedDocument, *, tenant_id: UUID
    ) -> KnowledgeDocumentView:
        with self._lock:
            catalog = self._catalog()
            records = self._records(catalog, knowledge_base_id)
            now = datetime.now(UTC)
            document_id, version_id = uuid4(), uuid4()
            version = DocumentVersionView(
                id=version_id,
                document_id=document_id,
                version_number=1,
                parsing_status="succeeded",
                chunking_status="succeeded",
                index_status="processing",
                chunking_strategy="local-bge-m3-v1",
                created_at=now,
            )
            document = KnowledgeDocumentView(
                id=document_id,
                knowledge_base_id=knowledge_base_id,
                original_filename=upload.filename,
                media_type=upload.media_type,
                status="published",
                current_version_id=version_id,
                created_at=now,
                updated_at=now,
                verified_at=upload.verified_at,
                review_due_at=upload.review_due_at,
                source_tiers=list(upload.source_tiers),
                source_urls=list(upload.source_urls),
                versions=[version],
            )
            relative_path = Path("documents") / str(document_id) / f"{version_id}.bin"
            self._write_bytes(relative_path, upload.content)
            record: dict[str, object] = {
                "view": document.model_dump(mode="json"),
                "content_path": relative_path.as_posix(),
                "sha256": hashlib.sha256(upload.content).hexdigest(),
                "chunks": [],
            }
            records.append(record)
            self._save(catalog)
            self._index_record(record, tenant_id=tenant_id)
            self._save(catalog)
            return KnowledgeDocumentView.model_validate(record["view"])

    def list_documents(
        self, knowledge_base_id: UUID, *, tenant_id: UUID
    ) -> KnowledgeDocumentListResponse:
        with self._lock:
            return KnowledgeDocumentListResponse(
                items=[
                    KnowledgeDocumentView.model_validate(item["view"])
                    for item in self._records(self._catalog(), knowledge_base_id)
                ]
            )

    def list_versions(self, document_id: UUID, *, tenant_id: UUID) -> DocumentVersionListResponse:
        with self._lock:
            view = KnowledgeDocumentView.model_validate(
                self._document(self._catalog(), document_id)["view"]
            )
            return DocumentVersionListResponse(document=view, items=view.versions)

    def list_chunks(
        self, document_id: UUID, version_id: UUID, *, tenant_id: UUID
    ) -> DocumentChunkListResponse:
        with self._lock:
            record = self._document(self._catalog(), document_id)
            document = KnowledgeDocumentView.model_validate(record["view"])
            if document.current_version_id != version_id:
                raise KnowledgeNotFound
            items: list[KnowledgeChunkView] = []
            for ordinal, chunk in enumerate(self._chunks(record)):
                text = str(chunk.get("text", "")).strip()
                digest = str(chunk.get("sha256", ""))
                if not text:
                    continue
                items.append(
                    KnowledgeChunkView(
                        id=UUID(str(chunk["id"])),
                        ordinal=ordinal,
                        offset=int(chunk.get("offset", 0)),
                        length=len(text),
                        text=text,
                        integrity_sha256=digest,
                    )
                )
            return DocumentChunkListResponse(
                document_id=document.id,
                document_version_id=version_id,
                items=items,
            )

    def publish(
        self, document_id: UUID, version_id: UUID, *, tenant_id: UUID
    ) -> LifecycleOperationResponse:
        self._require_version(document_id, version_id)
        return LifecycleOperationResponse(
            operation="publish",
            status="completed",
            document_id=document_id,
            version_id=version_id,
        )

    def rollback(
        self, document_id: UUID, version_id: UUID, *, tenant_id: UUID
    ) -> LifecycleOperationResponse:
        self._require_version(document_id, version_id)
        return LifecycleOperationResponse(
            operation="rollback",
            status="completed",
            document_id=document_id,
            version_id=version_id,
        )

    def delete(self, document_id: UUID, *, tenant_id: UUID) -> LifecycleOperationResponse:
        with self._lock:
            catalog = self._catalog()
            for records in catalog["documents"].values():
                for index, item in enumerate(records):
                    if item["view"]["id"] == str(document_id):
                        view = KnowledgeDocumentView.model_validate(item["view"])
                        self._delete_vectors(view.current_version_id)
                        records.pop(index)
                        self._save(catalog)
                        shutil.rmtree(
                            self._root / "documents" / str(document_id),
                            ignore_errors=True,
                        )
                        return LifecycleOperationResponse(
                            operation="delete",
                            status="completed",
                            document_id=document_id,
                        )
        raise KnowledgeNotFound

    def rebuild(
        self, document_id: UUID, version_id: UUID, *, tenant_id: UUID
    ) -> LifecycleOperationResponse:
        with self._lock:
            catalog = self._catalog()
            record = self._document(catalog, document_id)
            view = KnowledgeDocumentView.model_validate(record["view"])
            if view.current_version_id != version_id:
                raise KnowledgeNotFound
            self._index_record(record, tenant_id=tenant_id, rebuild=True)
            self._save(catalog)
            return LifecycleOperationResponse(
                operation="rebuild",
                status="completed",
                document_id=document_id,
                version_id=version_id,
            )

    def retrieve(
        self, request: RetrievalRequest, *, tenant_id: UUID, principal_id: UUID
    ) -> RetrievalResponse:
        if contains_prompt_injection(request.query):
            return self._empty(request.query, reason="unsafe_content")
        with self._lock:
            catalog = self._catalog()
            records = [
                item
                for base_id in request.knowledge_base_ids
                for item in self._records(catalog, base_id)
            ]
            catalog_changed = False
            for record in records:
                if not self._chunks(record):
                    self._index_record(record, tenant_id=tenant_id)
                    catalog_changed = True
            if catalog_changed:
                self._save(catalog)
            stale_records = [record for record in records if self._is_stale_record(record)]
            fresh_records = [record for record in records if record not in stale_records]
            chunks = {UUID(str(chunk["id"])): chunk for chunk in self._chunk_records(fresh_records)}
            stale_chunks = {
                UUID(str(chunk["id"])): chunk for chunk in self._chunk_records(stale_records)
            }
            if not chunks:
                reason = (
                    "stale_evidence"
                    if self._bm25(request.query, stale_chunks.values())
                    else "insufficient_evidence"
                )
                return self._empty(request.query, reason=reason)
            degradations: list[DegradationView] = []
            bm25 = self._bm25(request.query, chunks.values())
            vector: dict[UUID, float] = {}
            try:
                vector = self._vector_search(
                    request.query,
                    request.knowledge_base_ids,
                    tenant_id,
                    request.limit * 4,
                )
            except LocalRagUnavailable as error:
                degradations.append(
                    DegradationView(
                        kind="vector_degraded",
                        error_category="local_unavailable",
                        message=str(error),
                    )
                )
            candidates = self._fuse(
                bm25,
                {chunk_id: score for chunk_id, score in vector.items() if chunk_id in chunks},
                limit=request.limit * 4,
            )
            if not candidates:
                reason = (
                    "stale_evidence"
                    if self._bm25(request.query, stale_chunks.values())
                    else "insufficient_evidence"
                )
                return self._empty(request.query, degradations, reason=reason)
            reranked: dict[UUID, float] = {}
            try:
                reranked = self._rerank(request.query, [chunks[item] for item in candidates])
            except LocalRagUnavailable as error:
                degradations.append(
                    DegradationView(
                        kind="reranker_degraded",
                        error_category="local_unavailable",
                        message=str(error),
                    )
                )
            candidates = self._admit_candidates(bm25, candidates, reranked)
            if not candidates:
                reason = (
                    "stale_evidence"
                    if self._bm25(request.query, stale_chunks.values())
                    else "insufficient_evidence"
                )
                return self._empty(request.query, degradations, reason=reason)
            ordered = self._rank_candidates(
                request.query, candidates, reranked, chunks, limit=request.limit
            )
            hits = [
                self._hit(
                    chunks[chunk_id],
                    bm25.get(chunk_id),
                    vector.get(chunk_id),
                    candidates[chunk_id][2],
                    reranked.get(chunk_id),
                )
                for chunk_id in ordered
            ]
            evidence_ids = self._primary_evidence_ids(ordered, chunks)
            evidence_hits = [hit for hit in hits if hit.chunk_id in evidence_ids]
            if any(contains_prompt_injection(hit.excerpt) for hit in hits):
                return self._empty(
                    request.query,
                    degradations,
                    reason="unsafe_content",
                )
            conflicting_hits = self._conflicting_hits(hits)
            if conflicting_hits:
                return RetrievalResponse(
                    query=request.query,
                    answer=None,
                    refusal_reason="conflicting_evidence",
                    hits=conflicting_hits,
                    citations=[self._citation(hit) for hit in conflicting_hits],
                    degradations=degradations,
                )
            return RetrievalResponse(
                query=request.query,
                answer=self._extractive_retrieval_answer(evidence_hits),
                refusal_reason=None,
                hits=hits,
                citations=[self._citation(hit) for hit in evidence_hits],
                degradations=degradations,
            )

    def evaluate(
        self,
        request: EvaluationRequest,
        *,
        tenant_id: UUID,
        principal_id: UUID,
    ) -> EvaluationResponse:
        dataset = self._load_evaluation_dataset(request.dataset_id)
        selected = EvaluationDataset(
            dataset_id=dataset.dataset_id,
            version=dataset.version,
            cases=dataset.cases[: request.max_cases],
            digest_sha256=dataset.digest_sha256,
        )
        with self._lock:
            catalog = self._catalog()
            records = [
                item
                for base_id in request.knowledge_base_ids
                for item in self._records(catalog, base_id)
            ]
            catalog_changed = False
            for record in records:
                if not self._chunks(record):
                    self._index_record(record, tenant_id=tenant_id)
                    catalog_changed = True
            if catalog_changed:
                self._save(catalog)
            records = [record for record in records if not self._is_stale_record(record)]
            chunks = {UUID(str(item["id"])): item for item in self._chunk_records(records)}
            bm25_corpus = self._prepare_bm25(chunks.values())
        availability = {"vector": True, "reranker": True}
        observations = tuple(
            self._evaluate_case(
                case.query,
                case.case_id,
                request.knowledge_base_ids,
                chunks,
                bm25_corpus,
                availability,
                tenant_id=tenant_id,
            )
            for case in selected.cases
        )
        report = evaluate_dataset(selected, observations, k=_EVALUATION_K).to_dict()
        quality = report["quality"]
        operations = report["operations"]
        metrics: dict[str, int | float] = {
            **quality,
            "latency_p50_ms": operations["latency_ms"]["p50"],
            "latency_p95_ms": operations["latency_ms"]["p95"],
            "call_count": operations["call_count"],
            "failed_call_count": operations["failed_call_count"],
            "failure_rate": operations["failure_rate"],
            "estimated_cost_usd": operations["estimated_cost_usd"],
        }
        quality_gate_passed = bool(selected.cases) and all(
            float(metrics[name]) >= threshold for name, threshold in _QUALITY_THRESHOLDS.items()
        )
        quality_gate_passed = (
            quality_gate_passed and float(metrics["failure_rate"]) <= _MAX_EVALUATION_FAILURE_RATE
        )
        return EvaluationResponse(
            dataset_id=selected.dataset_id,
            dataset_version=selected.version,
            dataset_sha256=selected.digest_sha256,
            case_count=len(selected.cases),
            metrics=metrics,
            thresholds={
                **_QUALITY_THRESHOLDS,
                "max_failure_rate": _MAX_EVALUATION_FAILURE_RATE,
            },
            case_results=[
                self._evaluation_case_result(case, observation)
                for case, observation in zip(selected.cases, observations, strict=True)
            ],
            quality_gate_passed=quality_gate_passed,
        )

    @staticmethod
    def _evaluation_case_result(
        case: EvaluationCase,
        observation: EvaluationObservation,
    ) -> EvaluationCaseResultView:
        expected_documents = set(case.relevance)
        retrieved = tuple(observation.reranked_ids[:_EVALUATION_K])
        cited = set(observation.cited_ids)
        expected_citations = set(case.expected_citation_ids)
        recall: float | None = None
        reciprocal_rank: float | None = None
        citation_precision: float | None = None
        expected_citation_recall: float | None = None
        extractive_faithfulness: float | None = None
        reasons: list[str] = []
        if case.expected_refusal:
            if not observation.refused:
                reasons.append("missing_expected_refusal")
            if cited:
                reasons.append("unexpected_citation_on_refusal")
        else:
            recall = len(expected_documents & set(retrieved)) / len(expected_documents)
            reciprocal_rank = next(
                (
                    1.0 / rank
                    for rank, document_id in enumerate(retrieved, start=1)
                    if document_id in expected_documents
                ),
                0.0,
            )
            citation_precision = len(expected_documents & cited) / len(cited) if cited else 0.0
            expected_citation_recall = (
                len(expected_citations & cited) / len(expected_citations)
                if expected_citations
                else 1.0
            )
            normalized_evidence = tuple(
                " ".join(text.casefold().split()) for text in observation.cited_evidence_texts
            )
            extractive_faithfulness = (
                sum(
                    any(
                        " ".join(claim.casefold().split()) in evidence
                        for evidence in normalized_evidence
                    )
                    for claim in observation.answer_claims
                )
                / len(observation.answer_claims)
                if observation.answer_claims
                else 0.0
            )
            if observation.refused:
                reasons.append("unexpected_refusal")
            if recall < 1.0:
                reasons.append("missing_relevant_document")
            if citation_precision < 1.0:
                reasons.append("unsupported_citation")
            if expected_citation_recall < 1.0:
                reasons.append("missing_expected_citation")
            if extractive_faithfulness < 1.0:
                reasons.append("unsupported_answer_claim")
        if observation.failed_call_count:
            reasons.append("retrieval_dependency_failure")
        return EvaluationCaseResultView(
            case_id=case.case_id,
            language=case.language,
            query=case.query,
            expected_document_ids=list(case.relevance),
            baseline_document_ids=list(observation.baseline_ids),
            retrieved_document_ids=list(retrieved),
            cited_document_ids=list(observation.cited_ids),
            expected_refusal=case.expected_refusal,
            actual_refusal=observation.refused,
            recall_at_k=recall,
            reciprocal_rank=reciprocal_rank,
            citation_precision=citation_precision,
            expected_citation_recall=expected_citation_recall,
            extractive_faithfulness=extractive_faithfulness,
            latency_ms=observation.latency_ms,
            failed_call_count=observation.failed_call_count,
            passed=not reasons,
            failure_reasons=reasons,
        )

    def _load_evaluation_dataset(self, dataset_id: str) -> EvaluationDataset:
        root_source = self._evaluation_root
        candidate: Path | None = None
        try:
            if root_source.is_symlink():
                raise KnowledgeInputRejected("evaluation dataset path is invalid")
            root = root_source.resolve(strict=True)
            candidate = root / f"{dataset_id}.json"
            if candidate.is_symlink():
                raise KnowledgeInputRejected("evaluation dataset path is invalid")
            path = candidate.resolve(strict=True)
        except OSError as error:
            raise KnowledgeInputRejected("evaluation dataset is unavailable") from error
        if not root.is_dir() or path.parent != root:
            raise KnowledgeInputRejected("evaluation dataset path is invalid")
        try:
            dataset = load_evaluation_dataset(path)
        except (OSError, TypeError, ValueError) as error:
            raise KnowledgeInputRejected("evaluation dataset is invalid") from error
        if dataset.dataset_id != dataset_id:
            raise KnowledgeInputRejected("evaluation dataset identifier does not match")
        return dataset

    def _evaluate_case(
        self,
        query: str,
        case_id: str,
        base_ids: Sequence[UUID],
        chunks: dict[UUID, dict[str, object]],
        bm25_corpus: _LocalBm25Corpus,
        availability: dict[str, bool],
        *,
        tenant_id: UUID,
    ) -> EvaluationObservation:
        started = time.perf_counter()
        bm25 = self._score_bm25(query, bm25_corpus)
        baseline_ids = self._ranked_document_ids(
            sorted(bm25, key=lambda item: (-bm25[item], str(item))),
            chunks,
            limit=_EVALUATION_K,
        )
        vector: dict[UUID, float] = {}
        call_count = 0
        failed_call_count = 0
        if chunks and availability["vector"]:
            call_count += 1
            try:
                vector = self._vector_search(
                    query,
                    base_ids,
                    tenant_id,
                    _EVALUATION_K * 4,
                )
            except LocalRagUnavailable:
                failed_call_count += 1
        candidates = self._fuse(
            bm25,
            {chunk_id: score for chunk_id, score in vector.items() if chunk_id in chunks},
            limit=_EVALUATION_K * 4,
        )
        reranked: dict[UUID, float] = {}
        if candidates and availability["reranker"]:
            call_count += 1
            try:
                reranked = self._rerank(
                    query,
                    [chunks[item] for item in candidates],
                )
            except LocalRagUnavailable:
                failed_call_count += 1
        candidates = self._admit_candidates(bm25, candidates, reranked)
        ordered = self._rank_candidates(
            query, candidates, reranked, chunks, limit=_EVALUATION_K
        )
        evidence_ordered = self._primary_evidence_ids(ordered, chunks)
        reranked_ids = self._ranked_document_ids(
            ordered,
            chunks,
            limit=_EVALUATION_K,
        )
        cited_ids = self._ranked_document_ids(
            evidence_ordered,
            chunks,
            limit=_EVALUATION_K,
        )
        evidence_texts = tuple(str(chunks[chunk_id]["text"]) for chunk_id in evidence_ordered)
        return EvaluationObservation(
            case_id=case_id,
            baseline_ids=baseline_ids,
            reranked_ids=reranked_ids,
            cited_ids=cited_ids,
            refused=not candidates,
            latency_ms=(time.perf_counter() - started) * 1_000,
            estimated_cost_usd=0.0,
            call_count=call_count,
            failed_call_count=failed_call_count,
            answer_claims=evidence_texts,
            cited_evidence_texts=evidence_texts,
        )

    @staticmethod
    def _ranked_document_ids(
        ranked_chunk_ids: Iterable[UUID],
        chunks: dict[UUID, dict[str, object]],
        *,
        limit: int,
    ) -> tuple[str, ...]:
        result: list[str] = []
        for chunk_id in ranked_chunk_ids:
            document = chunks[chunk_id].get("document")
            if not isinstance(document, KnowledgeDocumentView):
                continue
            if document.original_filename not in result:
                result.append(document.original_filename)
            if len(result) >= limit:
                break
        return tuple(result)

    @staticmethod
    def _extractive_retrieval_answer(hits: Sequence[RetrievalHitView]) -> str:
        prefix = "以下内容直接摘自本地知识库，仅作为分析依据，不能授权工具执行：\n"
        parts: list[str] = []
        remaining = 4_096 - len(prefix)
        for index, hit in enumerate(hits, start=1):
            item = f"[{index}] {hit.excerpt.strip()}"
            separator = 1 if parts else 0
            if len(item) + separator > remaining:
                break
            parts.append(item)
            remaining -= len(item) + separator
        return prefix + "\n".join(parts)

    @staticmethod
    def _conflicting_hits(
        hits: Sequence[RetrievalHitView],
    ) -> list[RetrievalHitView]:
        by_claim: dict[str, dict[str, list[RetrievalHitView]]] = {}
        for hit in hits:
            for claim_id, stance in _CLAIM_MARKER.findall(hit.excerpt):
                claim = by_claim.setdefault(claim_id.casefold(), {})
                claim.setdefault(stance.casefold(), []).append(hit)
        conflicting_ids = {
            claim_id
            for claim_id, stances in by_claim.items()
            if stances.get("supports") and stances.get("counters")
        }
        if not conflicting_ids:
            return []
        selected: set[UUID] = set()
        for claim_id in conflicting_ids:
            for stance in ("supports", "counters"):
                selected.update(hit.chunk_id for hit in by_claim[claim_id][stance])
        return [hit for hit in hits if hit.chunk_id in selected]

    def _index_record(
        self, record: dict[str, object], *, tenant_id: UUID, rebuild: bool = False
    ) -> None:
        view = KnowledgeDocumentView.model_validate(record["view"])
        version = self._current_version(view)
        chunks: list[dict[str, object]] = []
        strategy = "rule-degraded-v1"
        failure_category: str | None = "unavailable"
        try:
            content = self._extract_text(record, view)
            chunks, strategy, failure_category = self._make_chunks(content, view)
            if not chunks:
                raise LocalRagUnavailable("No usable text content for indexing")
            # Persist lexical chunks before optional vector work. A stopped local
            # Embedding/Milvus stack must degrade to BM25, not erase RAG evidence.
            record["chunks"] = chunks
            vectors = self._embed([str(chunk["text"]) for chunk in chunks])
            self._delete_vectors(version.id)
            self._upsert_vectors(chunks, view, version.id, tenant_id, vectors)
            self._replace_version(
                record,
                version.id,
                index_status="succeeded",
                chunking_status="succeeded",
                chunking_strategy=strategy,
                chunking_failure_category=failure_category,
            )
        except LocalRagUnavailable:
            if chunks:
                self._replace_version(
                    record,
                    version.id,
                    index_status="failed",
                    chunking_status="succeeded",
                    chunking_strategy=strategy,
                    chunking_failure_category=failure_category,
                )
                return
            # Original bytes remain durable when text extraction itself failed.
            self._replace_version(
                record, version.id, index_status="failed", chunking_status="succeeded"
            )
            if rebuild:
                raise

    def _embed(self, texts: Sequence[str]) -> list[list[float]]:
        data: list[object] = []
        for start in range(0, len(texts), _EMBED_BATCH_SIZE):
            batch = list(texts[start : start + _EMBED_BATCH_SIZE])
            payload = self._model_request(
                "/v1/embeddings", {"model": "BAAI/bge-m3", "input": batch}
            )
            items = payload.get("data")
            if not isinstance(items, list) or len(items) != len(batch):
                raise LocalRagUnavailable(
                    "\u672c\u5730 Embedding \u670d\u52a1\u8fd4\u56de\u4e86\u65e0\u6548\u6570\u636e"
                )
            data.extend(items)
        try:
            vectors = [[float(value) for value in item["embedding"]] for item in data]
        except (KeyError, TypeError, ValueError) as error:
            raise LocalRagUnavailable(
                "\u672c\u5730 Embedding \u670d\u52a1\u8fd4\u56de\u4e86\u65e0\u6548\u5411\u91cf"
            ) from error
        if any(len(vector) != 1024 for vector in vectors):
            raise LocalRagUnavailable(
                "\u672c\u5730 Embedding \u5411\u91cf\u7ef4\u5ea6\u4e0d\u662f 1024"
            )
        return vectors

    @staticmethod
    def _admit_candidates(
        bm25: dict[UUID, float],
        candidates: dict[UUID, tuple[float | None, float | None, float]],
        reranked: dict[UUID, float],
    ) -> dict[UUID, tuple[float | None, float | None, float]]:
        """Reject unsupported vector neighbours instead of citing the nearest unrelated text."""
        if not candidates:
            return {}
        if bm25:
            return candidates
        if not reranked:
            return {}
        maximum = max(reranked.values(), default=0.0)
        threshold = max(_MIN_RERANK_RELEVANCE, maximum * _RELATIVE_RERANK_RELEVANCE)
        return candidates if maximum >= threshold else {}

    @classmethod
    def _rank_candidates(
        cls,
        query: str,
        candidates: dict[UUID, tuple[float | None, float | None, float]],
        reranked: dict[UUID, float],
        chunks: dict[UUID, dict[str, object]],
        *,
        limit: int,
    ) -> list[UUID]:
        ranked = sorted(
            candidates,
            key=lambda identifier: (
                -cls._title_affinity(query, chunks[identifier]),
                -reranked.get(identifier, -1.0),
                -candidates[identifier][2],
                str(identifier),
            ),
        )
        diverse: list[UUID] = []
        remaining: list[UUID] = []
        seen_documents: set[UUID] = set()
        for identifier in ranked:
            document = chunks[identifier].get("document")
            document_id = document.id if isinstance(document, KnowledgeDocumentView) else None
            if document_id is not None and document_id not in seen_documents:
                seen_documents.add(document_id)
                diverse.append(identifier)
            else:
                remaining.append(identifier)
        return (diverse + remaining)[:limit]

    @classmethod
    def _title_affinity(cls, query: str, chunk: dict[str, object]) -> int:
        document = chunk.get("document")
        if not isinstance(document, KnowledgeDocumentView):
            return 0
        query_terms = set(cls._lexical_terms(query)) | set(
            re.findall(r"[a-z0-9]+", query.casefold())
        )
        title_terms = set(cls._lexical_terms(document.original_filename)) | set(
            re.findall(r"[a-z0-9]+", document.original_filename.casefold())
        )
        return len(
            {
                term
                for term in query_terms & title_terms
                if len(term) >= 2 and term not in _GENERIC_TITLE_TERMS
            }
        )

    @staticmethod
    def _primary_evidence_ids(
        ordered: Sequence[UUID], chunks: dict[UUID, dict[str, object]]
    ) -> list[UUID]:
        if not ordered:
            return []
        first = chunks[ordered[0]].get("document")
        if not isinstance(first, KnowledgeDocumentView):
            return [ordered[0]]
        return [
            identifier
            for identifier in ordered
            if isinstance(chunks[identifier].get("document"), KnowledgeDocumentView)
            and chunks[identifier]["document"].id == first.id
        ]

    def _rerank(self, query: str, chunks: Sequence[dict[str, object]]) -> dict[UUID, float]:
        if not chunks:
            return {}
        payload = self._model_request(
            "/v1/rerank",
            {
                "model": "BAAI/bge-reranker-v2-m3",
                "query": query,
                "documents": [str(item["text"]) for item in chunks],
            },
        )
        data = payload.get("data")
        if not isinstance(data, list) or len(data) != len(chunks):
            raise LocalRagUnavailable(
                "\u672c\u5730\u91cd\u6392\u670d\u52a1\u8fd4\u56de\u4e86\u65e0\u6548\u6570\u636e"
            )
        try:
            return {
                UUID(str(chunks[int(item["index"])]["id"])): float(item["score"]) for item in data
            }
        except (KeyError, TypeError, ValueError, IndexError) as error:
            raise LocalRagUnavailable(
                "\u672c\u5730\u91cd\u6392\u670d\u52a1\u8fd4\u56de\u4e86\u65e0\u6548\u5206\u6570"
            ) from error

    def _model_request(self, path: str, payload: dict[str, object]) -> dict[str, object]:
        encoded = json.dumps(payload).encode("utf-8")
        headers = {"Content-Type": "application/json"}
        key = os.environ.get("SHIELDCHAIN_LOCAL_RAG_API_KEY", "")
        if key:
            headers["Authorization"] = f"Bearer {key}"
        try:
            with urlopen(
                Request(f"{self._model_url.rstrip('/')}{path}", encoded, headers),
                timeout=120,
            ) as response:
                value = json.loads(response.read().decode("utf-8"))
        except (HTTPError, URLError, OSError, ValueError) as error:
            raise LocalRagUnavailable(
                f"\u672c\u5730\u6a21\u578b\u670d\u52a1\u4e0d\u53ef\u7528\uff1a{error}"
            ) from error
        if not isinstance(value, dict):
            raise LocalRagUnavailable(
                "\u672c\u5730\u6a21\u578b\u670d\u52a1\u8fd4\u56de\u4e86\u65e0\u6548\u54cd\u5e94"
            )
        return value

    def _vector_search(
        self, query: str, base_ids: Sequence[UUID], tenant_id: UUID, limit: int
    ) -> dict[UUID, float]:
        vector = self._embed([query])[0]
        client = self._milvus_client()
        bases = ", ".join(json.dumps(str(value)) for value in base_ids)
        scope_filter = (
            f"tenant_id == {json.dumps(str(tenant_id))} and published == true "
            f"and knowledge_base_id in [{bases}]"
        )
        try:
            result = client.search(
                collection_name=self._collection,
                data=[vector],
                limit=min(limit, 100),
                filter=scope_filter,
                output_fields=[
                    "knowledge_base_id",
                    "document_id",
                    "document_version_id",
                ],
                search_params={"metric_type": "COSINE", "params": {}},
            )
            raw = result[0] if isinstance(result, list) and result else []
            return {
                UUID(str(item["id"])): max(0.0, min(1.0, (float(item["distance"]) + 1.0) / 2.0))
                for item in raw
            }
        except (KeyError, TypeError, ValueError, IndexError, Exception) as error:
            raise LocalRagUnavailable(
                f"Milvus \u5411\u91cf\u68c0\u7d22\u4e0d\u53ef\u7528\uff1a{error}"
            ) from error

    def _milvus_client(self) -> Any:
        try:
            from pymilvus import MilvusClient

            ensure_collection(uri=self._milvus_url, collection=self._collection)
            return MilvusClient(uri=self._milvus_url)
        except Exception as error:
            raise LocalRagUnavailable(f"Milvus \u4e0d\u53ef\u7528\uff1a{error}") from error

    def _upsert_vectors(
        self,
        chunks: Sequence[dict[str, object]],
        view: KnowledgeDocumentView,
        version_id: UUID,
        tenant_id: UUID,
        vectors: Sequence[Sequence[float]],
    ) -> None:
        client = self._milvus_client()
        rows = [
            {
                "id": str(chunk["id"]),
                "vector": list(vector),
                "tenant_id": str(tenant_id),
                "knowledge_base_id": str(view.knowledge_base_id),
                "document_id": str(view.id),
                "document_version_id": str(version_id),
                "sensitivity": "internal",
                "permission_tags": [],
                "published": True,
            }
            for chunk, vector in zip(chunks, vectors, strict=True)
        ]
        try:
            client.upsert(collection_name=self._collection, data=rows)
            client.flush(collection_name=self._collection)
        except Exception as error:
            raise LocalRagUnavailable(f"Milvus 鍚戦噺鍐欏叆涓嶅彲鐢細{error}") from error

    def _delete_vectors(self, version_id: UUID | None) -> None:
        if version_id is None:
            return
        try:
            self._milvus_client().delete(
                collection_name=self._collection,
                filter=f"document_version_id == {json.dumps(str(version_id))}",
            )
        except LocalRagUnavailable:
            # Deleting a local document must still delete its durable metadata and source bytes.
            pass

    def _make_chunks(
        self, content: str, view: KnowledgeDocumentView
    ) -> tuple[list[dict[str, object]], str, str | None]:
        cleaned = self._clean_text(content)
        if not cleaned:
            return [], "deepseek-semantic-v1", None
        try:
            segments = DeepSeekSemanticChunker.from_settings().chunk(cleaned)
            strategy, failure_category = "deepseek-semantic-v1", None
        except SemanticChunkingError:
            segments = self._rule_segments(cleaned)
            strategy, failure_category = "rule-degraded-v1", "unavailable"
        return self._chunks_for_segments(segments, view), strategy, failure_category

    @staticmethod
    def _rule_segments(content: str) -> tuple[SemanticSegment, ...]:
        segments: list[SemanticSegment] = []
        start = 0
        while start < len(content):
            end = min(len(content), start + _CHUNK_SIZE)
            if end < len(content):
                boundary = max(
                    content.rfind("\n", start + _CHUNK_SIZE // 2, end),
                    content.rfind("\u3002", start + _CHUNK_SIZE // 2, end),
                )
                if boundary > start:
                    end = boundary + 1
            text = content[start:end].strip()
            if text:
                segments.append(SemanticSegment(offset=start, text=text))
            if end >= len(content):
                break
            start = max(start + 1, end - _CHUNK_OVERLAP)
        return tuple(segments)

    def _chunks_for_segments(
        self, segments: Sequence[SemanticSegment], view: KnowledgeDocumentView
    ) -> list[dict[str, object]]:
        version = self._current_version(view)
        chunks: list[dict[str, object]] = []
        for segment in segments:
            digest = hashlib.sha256(segment.text.encode("utf-8")).hexdigest()
            chunk_id = uuid5(_CHUNK_NAMESPACE, f"{version.id}:{len(chunks)}:{digest}")
            chunks.append(
                {
                    "id": str(chunk_id),
                    "text": segment.text,
                    "offset": segment.offset,
                    "sha256": digest,
                }
            )
        return chunks

    @staticmethod
    def _clean_text(content: str) -> str:
        value = re.sub(r"```.*?```", " ", content, flags=re.DOTALL)
        value = re.sub(r"!?(?:\[[^\]]*\])\([^)]*\)", " ", value)
        value = re.sub(r"[*_`>#]", "", value)
        return re.sub(r"[ \t]+", " ", value).strip()

    def _extract_text(self, record: dict[str, object], view: KnowledgeDocumentView) -> str:
        raw = self._read_bytes(record["content_path"])
        suffix = Path(view.original_filename).suffix.casefold()
        if suffix in {".txt", ".md", ".csv", ".html", ".htm"}:
            text = raw.decode("utf-8", errors="ignore")
            if suffix in {".html", ".htm"}:
                from bs4 import BeautifulSoup

                soup = BeautifulSoup(text, "html.parser")
                for node in soup(
                    ["script", "style", "noscript", "template", "iframe", "object", "embed"]
                ):
                    node.decompose()
                content = soup.select_one("#BodyLabel, .main-content, .side-right-column")
                return (content or soup).get_text("\n", strip=True)
            return text
        if suffix == ".pdf":
            from io import BytesIO

            from pypdf import PdfReader

            return "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(raw)).pages)
        if suffix == ".docx":
            from io import BytesIO

            from docx import Document

            return "\n".join(paragraph.text for paragraph in Document(BytesIO(raw)).paragraphs)
        if suffix == ".xlsx":
            from io import BytesIO

            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            return "\n".join(
                " ".join(str(cell) for cell in row if cell is not None)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows(values_only=True)
            )
        return raw.decode("utf-8", errors="ignore")

    def _bm25(self, query: str, chunks: Iterable[dict[str, object]]) -> dict[UUID, float]:
        return self._score_bm25(query, self._prepare_bm25(chunks))

    @staticmethod
    def _prepare_bm25(chunks: Iterable[dict[str, object]]) -> _LocalBm25Corpus:
        source = tuple(chunks)
        tokenized = tuple(
            Counter(LocalKnowledgeService._lexical_terms(str(chunk["text"]))) for chunk in source
        )
        lengths = tuple(sum(counter.values()) for counter in tokenized)
        average = sum(lengths) / len(lengths) if lengths else 1.0
        return _LocalBm25Corpus(
            entries=tuple(zip(source, tokenized, lengths, strict=True)),
            average_length=average or 1.0,
            document_frequency=Counter(term for counter in tokenized for term in set(counter)),
        )

    @staticmethod
    def _score_bm25(query: str, corpus: _LocalBm25Corpus) -> dict[UUID, float]:
        terms = LocalKnowledgeService._lexical_terms(query)
        if not corpus.entries or not terms:
            return {}
        scores: dict[UUID, float] = {}
        for chunk, counts, length in corpus.entries:
            score = 0.0
            for term in terms:
                frequency = counts[term]
                if frequency:
                    inverse = math.log(
                        1
                        + (len(corpus.entries) - corpus.document_frequency[term] + 0.5)
                        / (corpus.document_frequency[term] + 0.5)
                    )
                    score += (
                        inverse
                        * (frequency * 2.0)
                        / (frequency + 1.2 * (1 - 0.75 + 0.75 * length / corpus.average_length))
                    )
            if score:
                scores[UUID(str(chunk["id"]))] = score
        maximum = max(scores.values(), default=0.0)
        return {chunk_id: score / maximum for chunk_id, score in scores.items()} if maximum else {}

    @staticmethod
    def _lexical_terms(text: str) -> list[str]:
        """Keep CJK phrase meaning while retaining deterministic offline tokenization."""
        normalized = text.casefold()
        terms = _TERM.findall(normalized)
        for match in _CJK_SEQUENCE.finditer(normalized):
            sequence = match.group()
            for width in (2, 3):
                terms.extend(
                    sequence[index : index + width] for index in range(len(sequence) - width + 1)
                )
        return terms

    @staticmethod
    def _fuse(
        bm25: dict[UUID, float], vector: dict[UUID, float], *, limit: int
    ) -> dict[UUID, tuple[float | None, float | None, float]]:
        identifiers = set(bm25) | set(vector)
        fused = {
            identifier: (
                bm25.get(identifier),
                vector.get(identifier),
                0.45 * bm25.get(identifier, 0.0) + 0.55 * vector.get(identifier, 0.0),
            )
            for identifier in identifiers
        }
        return dict(sorted(fused.items(), key=lambda item: (-item[1][2], str(item[0])))[:limit])

    def _chunk_records(self, records: Iterable[dict[str, object]]) -> Iterable[dict[str, object]]:
        for record in records:
            document = KnowledgeDocumentView.model_validate(record["view"])
            for chunk in self._chunks(record):
                yield {
                    **chunk,
                    "document": document,
                    "document_sha256": str(record["sha256"]),
                }

    @staticmethod
    def _chunks(record: dict[str, object]) -> list[dict[str, object]]:
        chunks = record.get("chunks", [])
        return (
            [item for item in chunks if isinstance(item, dict)] if isinstance(chunks, list) else []
        )

    @staticmethod
    def _current_version(view: KnowledgeDocumentView) -> DocumentVersionView:
        for version in view.versions:
            if version.id == view.current_version_id:
                return version
        raise KnowledgeNotFound

    def _replace_version(
        self, record: dict[str, object], version_id: UUID, **changes: object
    ) -> None:
        view = KnowledgeDocumentView.model_validate(record["view"])
        versions = [
            version.model_copy(update=changes) if version.id == version_id else version
            for version in view.versions
        ]
        record["view"] = view.model_copy(
            update={"versions": versions, "updated_at": datetime.now(UTC)}
        ).model_dump(mode="json")

    def _hit(
        self,
        chunk: dict[str, object],
        bm25: float | None,
        vector: float | None,
        fusion: float,
        reranker: float | None,
    ) -> RetrievalHitView:
        document = chunk["document"]
        if not isinstance(document, KnowledgeDocumentView):
            raise TypeError("local chunk metadata is invalid")
        version = self._current_version(document)
        return RetrievalHitView(
            chunk_id=UUID(str(chunk["id"])),
            knowledge_base_id=document.knowledge_base_id,
            document_id=document.id,
            document_version_id=version.id,
            document_title=document.original_filename,
            excerpt=str(chunk["text"]),
            heading_path=[],
            structural_location=f"chunk@{chunk['offset']}",
            bm25_score=bm25,
            vector_score=vector,
            fusion_score=fusion,
            reranker_score=reranker,
            updated_at=document.updated_at,
            integrity_sha256=str(chunk["sha256"]),
            verified_at=document.verified_at,
            review_due_at=document.review_due_at,
            source_tiers=document.source_tiers,
            source_urls=document.source_urls,
        )

    @staticmethod
    def _citation(hit: RetrievalHitView) -> CitationView:
        return CitationView(citation_id=f"local:{hit.chunk_id}", **hit.model_dump())

    @staticmethod
    def _empty(
        query: str,
        degradations: list[DegradationView] | None = None,
        *,
        reason: str = "insufficient_evidence",
    ) -> RetrievalResponse:
        return RetrievalResponse(
            query=query,
            answer=None,
            refusal_reason=reason,
            hits=[],
            citations=[],
            degradations=degradations or [],
        )

    @staticmethod
    def _is_stale_record(record: dict[str, object]) -> bool:
        document = KnowledgeDocumentView.model_validate(record["view"])
        return bool(document.review_due_at and document.review_due_at < datetime.now(UTC).date())

    def _catalog(self) -> dict[str, Any]:
        if not self._catalog_path.exists():
            return {"bases": [], "documents": {}}
        value = json.loads(self._catalog_path.read_text(encoding="utf-8"))
        if (
            not isinstance(value, dict)
            or not isinstance(value.get("bases"), list)
            or not isinstance(value.get("documents"), dict)
        ):
            raise TypeError("local knowledge catalog is invalid")
        return value

    def _save(self, catalog: dict[str, Any]) -> None:
        with tempfile.NamedTemporaryFile(
            "w", encoding="utf-8", dir=self._root, delete=False
        ) as handle:
            json.dump(catalog, handle, ensure_ascii=False, indent=2, sort_keys=True)
            temporary = Path(handle.name)
        temporary.replace(self._catalog_path)

    def _records(self, catalog: dict[str, Any], knowledge_base_id: UUID) -> list[dict[str, object]]:
        if not any(item.get("id") == str(knowledge_base_id) for item in catalog["bases"]):
            raise KnowledgeNotFound
        records = catalog["documents"].get(str(knowledge_base_id))
        if not isinstance(records, list):
            raise TypeError("local knowledge records are invalid")
        return records

    def _document(self, catalog: dict[str, Any], document_id: UUID) -> dict[str, object]:
        for records in catalog["documents"].values():
            for item in records:
                if item["view"]["id"] == str(document_id):
                    return item
        raise KnowledgeNotFound

    def _require_version(self, document_id: UUID, version_id: UUID) -> None:
        view = KnowledgeDocumentView.model_validate(
            self._document(self._catalog(), document_id)["view"]
        )
        if all(version.id != version_id for version in view.versions):
            raise KnowledgeNotFound

    def _write_bytes(self, relative_path: Path, content: bytes) -> None:
        target = self._root / relative_path
        target.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile("wb", dir=target.parent, delete=False) as handle:
            handle.write(content)
            temporary = Path(handle.name)
        temporary.replace(target)

    def _read_bytes(self, relative_path: object) -> bytes:
        if not isinstance(relative_path, str):
            return b""
        try:
            return (self._root / relative_path).read_bytes()
        except OSError:
            return b""
